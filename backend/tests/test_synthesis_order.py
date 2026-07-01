"""Tests for vendor synthesis-order export — no network/account access, file generation only."""

from io import BytesIO

import pytest
from openpyxl import load_workbook

from ..formats import VENDORS, export_order
from ..models.sequence import MoleculeType, Sequence

HBB_SEQ = Sequence(
    id="HBB",
    name="HBB",
    seq="ATGGTGCACCTGACTCCTGAGGAGAAGTCTGCCGTTACTGCCCTGTGGGGCAAGGTGAACGTGGATGAA",
    molecule_type=MoleculeType.DNA,
    description="hemoglobin subunit beta",
)


def test_all_vendors_have_instructions() -> None:
    assert set(VENDORS) == {"genewiz", "idt", "twist", "eurofins"}
    for profile in VENDORS.values():
        assert profile.instructions
        assert profile.file_format in ("xlsx", "csv")


def test_export_unknown_vendor_raises() -> None:
    with pytest.raises(ValueError):
        export_order(HBB_SEQ, "not_a_real_vendor")


def test_export_empty_range_raises() -> None:
    with pytest.raises(ValueError):
        export_order(HBB_SEQ, "idt", start=5, end=5)


def test_export_genewiz_xlsx_columns() -> None:
    data, filename, instructions = export_order(HBB_SEQ, "genewiz")
    assert filename.endswith(".xlsx")
    assert instructions
    wb = load_workbook(BytesIO(data))
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == ["Sequence Name", "Sequence Type", "5' Flanking", "Sequence", "3' Flanking"]
    row = [c.value for c in ws[2]]
    assert row[0] == "HBB"
    assert row[3] == HBB_SEQ.seq


def test_export_idt_csv() -> None:
    data, filename, instructions = export_order(HBB_SEQ, "idt")
    assert filename.endswith(".csv")
    text = data.decode("utf-8")
    assert text.splitlines() == ["Name,Sequence", f"HBB,{HBB_SEQ.seq}"]


def test_export_sub_range() -> None:
    data, _, _ = export_order(HBB_SEQ, "idt", start=0, end=6)
    text = data.decode("utf-8")
    assert text.splitlines()[1] == "HBB,ATGGTG"


def test_eurofins_marked_unverified() -> None:
    assert VENDORS["eurofins"].verified is False
